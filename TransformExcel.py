import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from tkinter import Tk, filedialog, Button, Label, messagebox

def transform_data(input_file_path, output_file_path):
    df_input = pd.read_excel(input_file_path, header=5)
    df_input = df_input.dropna(subset=['Order No'])
    wb = Workbook()
    ws = wb.active

    # Cell colors
    header_fill = PatternFill(start_color="7CE086", end_color="7CE086", fill_type="solid")
    item_fill = PatternFill(start_color="7BA9E1", end_color="7BA9E1", fill_type="solid")
    expense_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

    # HEADER columns
    ws.append([
        "HEADER", "No Form", "Tgl Pesanan", "No Pelanggan", "No PO", "Alamat", "Kena PPN", 
        "Total Termasuk PPN", "Diskon Pesanan (%)", "Diskon Pesanan (Rp)", "Keterangan", 
        "Nama Cabang", "Pengiriman", "Tgl Pengiriman", "FOB", "Syarat Pembayaran"
    ])
    for cell in ws[1]:
        cell.fill = header_fill 

    # ITEM columns
    ws.append([
        "ITEM", "Kode Barang", "Nama Barang", "Kuantitas", "Satuan", "Harga Satuan", 
        "Diskon Barang (%)", "Diskon Barang (Rp)", "Catatan Barang", "Nama Dept Barang", 
        "No Proyek Barang", "Nama Gudang", "ID Salesman", "Kustom Karakter 1", 
        "Kustom Karakter 2", "Kustom Karakter 3"
    ])
    for cell in ws[2]:
        cell.fill = item_fill  

    # EXPENSE columns
    ws.append([
        "EXPENSE", "No Biaya", "Nama Biaya", "Nilai Biaya", "Catatan Biaya", "Nama Dept Biaya", 
        "No Proyek Biaya", "Kategori Keuangan 1", "Kategori Keuangan 2", "Kategori Keuangan 3", 
        "Kategori Keuangan 4", "Kategori Keuangan 5", "Kategori Keuangan 6", "Kategori Keuangan 7", 
        "Kategori Keuangan 8", "Kategori Keuangan 9"
    ])
    for cell in ws[3]:
        cell.fill = expense_fill 

    unique_orders = df_input['Order No'].unique()
    for order_no in unique_orders:
        order_data = df_input[df_input['Order No'] == order_no]

        # HEADER rows
        ws.append([
            "HEADER", 
            order_no,
            order_data.iloc[0]["Posted Date"], 
            order_data.iloc[0]["Customer Code"], 
            None, 
            order_data.iloc[0]["Address"], 
            None, None, None, None, None, None, None, None, None, None 
        ])
        ws[f"A{ws.max_row}"].fill = header_fill 

        # ITEM rows
        for _, row in order_data.iterrows():
            ws.append([
                "ITEM",
                row["Item Code"], 
                row["Item Name"], 
                row["Quantity"], 
                row["UOM"],  
                row["Unit Price"],  
                row["Discount"],  
                None, 
                None, None, None, None, 
                row["DSR Code"],  
                None, None, None  
            ])
            ws[f"A{ws.max_row}"].fill = item_fill 

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2 

    wb.save(output_file_path)

def select_files():
    file_paths = filedialog.askopenfilenames(title="Select Input Files", filetypes=[("Excel Files", "*.xlsx")])
    if file_paths:
        process_files(file_paths)

def process_files(file_paths):
    output_dir = "output"
    os.makedirs(output_dir, exist_ok=True)

    for input_file in file_paths:
        file_name = os.path.basename(input_file)
        output_file_name = f"{os.path.splitext(file_name)[0]}_output.xlsx"
        output_file_path = os.path.join(output_dir, output_file_name)
        
        transform_data(input_file, output_file_path)
    
    messagebox.showinfo("Success", "Files processed and saved in the 'output' folder.")

root = Tk()
root.title("Excel Data Processor")
root.geometry("300x150")

label = Label(root, text="Select files to process:")
label.pack(pady=10)

select_button = Button(root, text="Select Files", command=select_files)
select_button.pack(pady=5)

root.mainloop()
