import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from tkinter import Tk, filedialog, Button, Label, messagebox, Frame
import xml.etree.ElementTree as ET
from xml.dom import minidom

# Function to transform data to excel
def transform_data(input_file_path, output_file_path):
    df_input = pd.read_excel(input_file_path, header=5)
    df_input = df_input.dropna(subset=['Order No'])
    wb = Workbook()
    ws = wb.active

    # Color fills for headers
    header_fill = PatternFill(start_color="7CE086", end_color="7CE086", fill_type="solid")
    item_fill = PatternFill(start_color="7BA9E1", end_color="7BA9E1", fill_type="solid")
    expense_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

    # Add headers with color fills
    ws.append(["HEADER", "No Form", "Tgl Pesanan", "No Pelanggan", "No PO", "Alamat", "Kena PPN", "Total Termasuk PPN",
               "Diskon Pesanan (%)", "Diskon Pesanan (Rp)", "Keterangan", "Nama Cabang", "Pengiriman", 
               "Tgl Pengiriman", "FOB", "Syarat Pembayaran"])
    for cell in ws[1]:
        cell.fill = header_fill

    ws.append(["ITEM", "Kode Barang", "Nama Barang", "Kuantitas", "Satuan", "Harga Satuan", "Diskon Barang (%)", 
               "Diskon Barang (Rp)", "Catatan Barang", "Nama Dept Barang", "No Proyek Barang", "Nama Gudang", 
               "ID Salesman", "Kustom Karakter 1", "Kustom Karakter 2", "Kustom Karakter 3"])
    for cell in ws[2]:
        cell.fill = item_fill

    ws.append(["EXPENSE", "No Biaya", "Nama Biaya", "Nilai Biaya", "Catatan Biaya", "Nama Dept Biaya", 
               "No Proyek Biaya", "Kategori Keuangan 1", "Kategori Keuangan 2", "Kategori Keuangan 3", 
               "Kategori Keuangan 4", "Kategori Keuangan 5", "Kategori Keuangan 6", "Kategori Keuangan 7", 
               "Kategori Keuangan 8", "Kategori Keuangan 9"])
    for cell in ws[3]:
        cell.fill = expense_fill

    unique_orders = df_input['Order No'].unique()
    for order_no in unique_orders:
        order_data = df_input[df_input['Order No'] == order_no]

        ws.append(["HEADER", order_no, order_data.iloc[0]["Posted Date"], order_data.iloc[0]["Customer Code"], None, 
                   order_data.iloc[0]["Address"], None, None, None, None, None, None, None, None, None, None])
        ws[f"A{ws.max_row}"].fill = header_fill

        for _, row in order_data.iterrows():
            ws.append(["ITEM", row["Item Code"], row["Item Name"], row["Quantity"], row["UOM"], row["Unit Price"], 
                       row["Discount"], None, None, None, None, None, row["DSR Code"], None, None, None])
            ws[f"A{ws.max_row}"].fill = item_fill

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(output_file_path)

# Function to create XML for Accurate Desktop
def create_sales_order_xml(input_file_path, output_file_path):
    df = pd.read_excel(input_file_path, skiprows=5)
    df.columns = [col.strip() for col in df.columns]

    def safe_str(value):
        return '' if pd.isna(value) else str(value)

    df = df.dropna(subset=['Order No'])

    root = ET.Element("NMEXML", EximID="13", BranchCode="2040822216", ACCOUNTANTCOPYID="")
    transactions = ET.SubElement(root, "TRANSACTIONS", OnError="CONTINUE")
    unique_orders = df['Order No'].unique()

    for order_no in unique_orders:
        sales_order = ET.SubElement(transactions, "SALESORDER", operation="Add", REQUESTID="1")
        ET.SubElement(sales_order, "TRANSACTIONID").text = "" 

        order_df = df[df['Order No'] == order_no]
        key_id_counter = 0

        for _, row in order_df.iterrows():
            item_line = ET.SubElement(sales_order, "ITEMLINE", operation="Add")
            ET.SubElement(item_line, "KeyID").text = str(key_id_counter)
            ET.SubElement(item_line, "ITEMNO").text = safe_str(row.get('Item Code', ''))
            ET.SubElement(item_line, "QUANTITY").text = safe_str(row.get('Quantity', ''))
            ET.SubElement(item_line, "ITEMUNIT").text = safe_str(row.get('UOM', ''))
            ET.SubElement(item_line, "UNITRATIO").text = ""  
            ET.SubElement(item_line, "ITEMRESERVED1")
            ET.SubElement(item_line, "ITEMRESERVED2")
            ET.SubElement(item_line, "ITEMRESERVED3")
            ET.SubElement(item_line, "ITEMRESERVED4")
            ET.SubElement(item_line, "ITEMRESERVED5")
            ET.SubElement(item_line, "ITEMRESERVED6")
            ET.SubElement(item_line, "ITEMRESERVED7")
            ET.SubElement(item_line, "ITEMRESERVED8")
            ET.SubElement(item_line, "ITEMRESERVED9")
            ET.SubElement(item_line, "ITEMRESERVED10")
            ET.SubElement(item_line, "ITEMOVDESC").text = safe_str(row.get('Item Name', ''))
            ET.SubElement(item_line, "UNITPRICE").text = safe_str(row.get('Unit Price', ''))
            ET.SubElement(item_line, "DISCPC").text = "" 
            ET.SubElement(item_line, "TAXCODES").text = ""
            ET.SubElement(item_line, "GROUPSEQ").text = ""
            ET.SubElement(item_line, "QTYSHIPPED").text = ""
            
            key_id_counter += 1

        ET.SubElement(sales_order, "SONO").text = safe_str(order_no)
        ET.SubElement(sales_order, "SODATE").text = safe_str(order_df['Posted Date'].iloc[0] if not order_df.empty else "")
        ET.SubElement(sales_order, "TAX1ID").text = ""
        ET.SubElement(sales_order, "TAX1CODE").text = ""
        ET.SubElement(sales_order, "TAX2CODE")
        ET.SubElement(sales_order, "TAX1RATE").text = ""
        ET.SubElement(sales_order, "TAX2RATE").text = ""
        ET.SubElement(sales_order, "TAX1AMOUNT").text = ""
        ET.SubElement(sales_order, "TAX2AMOUNT").text = ""
        ET.SubElement(sales_order, "RATE").text = ""
        ET.SubElement(sales_order, "TAXINCLUSIVE").text = ""
        ET.SubElement(sales_order, "CUSTOMERISTAXABLE").text = ""
        ET.SubElement(sales_order, "CASHDISCOUNT").text = ""
        ET.SubElement(sales_order, "CASHDISCPC").text = ""
        ET.SubElement(sales_order, "FREIGHT").text = ""
        ET.SubElement(sales_order, "TERMSID").text = ""
        ET.SubElement(sales_order, "SHIPVIAID").text = ""
        ET.SubElement(sales_order, "FOB").text = ""
        ET.SubElement(sales_order, "ESTSHIPDATE").text = ""
        ET.SubElement(sales_order, "DESCRIPTION").text = ""
        ET.SubElement(sales_order, "SHIPTO1").text = safe_str(order_df['Address'].iloc[0] if not order_df.empty else "")
        ET.SubElement(sales_order, "SHIPTO2").text = safe_str(order_df['District'].iloc[0] if not order_df.empty else "")
        ET.SubElement(sales_order, "SHIPTO3").text = ""
        ET.SubElement(sales_order, "SHIPTO4").text = ""
        ET.SubElement(sales_order, "SHIPTO5").text = ""
        ET.SubElement(sales_order, "DP").text = ""
        ET.SubElement(sales_order, "DPACCOUNTID").text = ""
        ET.SubElement(sales_order, "DPUSED").text = ""
        ET.SubElement(sales_order, "CUSTOMERID").text = safe_str(order_df['Customer Code'].iloc[0] if not order_df.empty else "")
        ET.SubElement(sales_order, "PONO").text = ""
        ET.SubElement(sales_order, "CURRENCYNAME").text = "IDR"
        
    xml_str = minidom.parseString(ET.tostring(root)).toprettyxml(indent="    ")
    with open(output_file_path, "w", encoding="utf-8") as f:
        f.write(xml_str)

# UI functions for file selection and processing
def select_files(file_type):
    file_paths = filedialog.askopenfilenames(title="Select Input Files", filetypes=[("Excel Files", "*.xlsx")])
    if file_paths:
        process_files(file_paths, file_type)

def process_files(file_paths, file_type):
    output_dir = "output"
    os.makedirs(output_dir, exist_ok=True)

    for input_file in file_paths:
        file_name = os.path.basename(input_file)
        if file_type == "excel":
            output_file_name = f"{os.path.splitext(file_name)[0]}_AO.xlsx"
            output_file_path = os.path.join(output_dir, output_file_name)
            transform_data(input_file, output_file_path)
        elif file_type == "xml":
            output_file_name = f"{os.path.splitext(file_name)[0]}_AD.xml"
            output_file_path = os.path.join(output_dir, output_file_name)
            create_sales_order_xml(input_file, output_file_path)

    messagebox.showinfo("Success", f"Files processed and saved in the '{output_dir}' folder.")

# Creating the GUI
root = Tk()
root.title("Data Processor")
root.geometry("400x300")
root.configure(bg="#e6f5ff")

frame = Frame(root, bg="#cfe2f3", bd=2, relief="solid")
frame.pack(pady=20, padx=20, fill="both", expand=True)

Label(frame, text="Data Processing Tool", font=("Arial", 16, "bold"), bg="#cfe2f3").pack(pady=10)
Label(frame, text="Select the output type for data conversion:", font=("Arial", 10), bg="#cfe2f3").pack(pady=5)

Button(frame, text="Excel Format (Accurate Online)", command=lambda: select_files("excel"),
       font=("Arial", 12), bg="#7CE086", fg="white", relief="groove", width=25).pack(pady=10)
Button(frame, text="XML Format (Accurate Desktop)", command=lambda: select_files("xml"),
       font=("Arial", 12), bg="#7BA9E1", fg="white", relief="groove", width=25).pack(pady=10)

Label(frame, text="Output files are saved in the 'output' folder.", font=("Arial", 8, "italic"), bg="#cfe2f3").pack(pady=10)

root.mainloop()
